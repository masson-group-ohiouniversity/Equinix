"""equinix_io.py"""
import re
import io
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from datetime import datetime
from PIL import Image, ImageDraw, ImageFont
from equinix_network import compute_variable_curve
from equinix_kinetics import _collect_all_kinetic_species
from equinix_curve import convert_exp_x

__all__ = ['export_to_excel', 'generate_parameters_text', 'generate_kinetics_parameters_text', 'text_to_image', 'create_snapshot', '_export_kinetics_excel', 'load_experimental_data', 'load_spectra_data', '_pub_tight_bounds', '_pub_figure_bytes', '_pub_axis_range', '_pub_download_button', '_plot_backcalc_dots', '_infer_unit', '_infer_y_label',
           'Image', 'ImageDraw', 'ImageFont']


def export_to_excel(curve, x_vals, parsed, params, network, script_text, logK_vals,
                    script_path=None, input_path=None):
    """
    Export titration data to timestamped Excel file with 3 tabs.
    
    Tabs:
    1. "data": Simulation results (volume, x-axis, species, variables)  
    2. "script": Original input script (one line per row)
    3. "parameters": Current parameter values in script format (includes fitted values)
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # Generate timestamp filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Equinix_{timestamp}.xlsx"
    
    # Find the concentration key that matches the primary component
    primary_component = params["primary_component"]
    primary_conc_key = None
    for cname in parsed["concentrations"].keys():
        root = cname[:-1] if cname.endswith("0") else cname
        if root == primary_component:
            primary_conc_key = cname
            break
    if primary_conc_key is None:
        primary_conc_key = list(parsed["concentrations"].keys())[0]
    
    n_primary = parsed["concentrations"][primary_conc_key] * params["V0_mL"] / 1000
    tit_mM    = parsed["titrant"][params["titrant_key"]]
    is_solid  = params.get("titrant_is_solid", False)

    # ── TAB 1: DATA ───────────────────────────────────────────────────────────
    x_col_name = parsed.get("plot_x_expr", "X-axis") or "X-axis"
    if is_solid:
        data = {x_col_name: x_vals}
    else:
        n_points = len(x_vals)
        xs       = np.linspace(0, params["maxEquiv"], n_points)
        V_add_mL = []
        for eq in xs:
            n_tit   = eq * n_primary
            V_add_L = n_tit / max(tit_mM, 1e-12)
            V_add_mL.append(V_add_L * 1000)
        data = {'V_add(mL)': V_add_mL, x_col_name: x_vals}

    for species in network["all_species"]:
        if species in curve:
            data[species] = curve[species]

    variables = parsed.get("variables", {})
    for var_name, expression in variables.items():
        try:
            data[var_name] = compute_variable_curve(var_name, variables, curve, network, x_vals)
        except Exception:
            pass

    df_data = pd.DataFrame(data)

    # ── TAB 2: SCRIPT ─────────────────────────────────────────────────────────
    script_lines = script_text.split('\n')
    df_script = pd.DataFrame({'Script': script_lines})

    # ── TAB 3: PARAMETERS ─────────────────────────────────────────────────────
    def _param_block(section):
        """Return rebuilt lines for a given section (header + content, no blank separator)."""
        lines = []
        if section in ("concentrations",):
            lines.append("$concentrations")
            for cname, cval in parsed["concentrations"].items():
                current_val = params["conc0"].get(cname[:-1] if cname.endswith("0") else cname, cval)
                lines.append(f"{cname} = {current_val:.4f}")
        elif section == "volumes":
            lines.append("$volumes")
            lines.append(f"V0 = {params['V0_mL']:.4f}")
        elif section == "titrant":
            if params.get("titrant_is_solid", False):
                lines.append("$titrant solid")
                tit_ratios = parsed.get("titrant_ratios", {})
                for tkey in parsed["titrant"].keys():
                    lines.append(f"{tkey}; {tit_ratios.get(tkey, 1.0):.4f}")
            else:
                lines.append("$titrant")
                for tkey, tval in parsed["titrant"].items():
                    current_val = params["titrant_mMs"].get(tkey[:-1] if tkey.endswith("t") else tkey, tval)
                    lines.append(f"{tkey} = {current_val:.4f}")
        elif section in ("reactions", "equilibria"):
            lines.append("$reactions")
            for eq in parsed["equilibria"]:
                reactants_str = []
                for coeff, species in eq["reactants"]:
                    reactants_str.append(f"{coeff} {species}" if coeff > 1 else species)
                products_str = []
                if "products" in eq:
                    for coeff, species in eq["products"]:
                        products_str.append(f"{coeff} {species}" if coeff > 1 else species)
                elif "product" in eq:
                    prod_coeff, prod_species = eq["product"]
                    products_str.append(f"{prod_coeff} {prod_species}" if prod_coeff > 1 else prod_species)
                kname        = eq["kname"]
                current_logK = logK_vals.get(kname, eq["logK"])
                lines.append(f"{' + '.join(reactants_str)} = {' + '.join(products_str)}; {kname} = {current_logK:.4f}")
        elif section == "variables" and parsed.get("variables"):
            lines.append("$variables")
            for var_name, expression in parsed["variables"].items():
                lines.append(f"{var_name} = {expression}")
        elif section == "plot" and (parsed.get("plot_y") or parsed.get("plot_x_expr")):
            lines.append("$plot")
            lines.append(f"xmax = {params['maxEquiv']:.4f}")
            if parsed.get("plot_x_expr"):
                lines.append(f"x = {parsed['plot_x_expr']}")
            if parsed.get("plot_y"):
                lines.append(f"y = {', '.join(parsed['plot_y'])}")
        elif section == "nmr" and parsed.get("nmr"):
            nmr = parsed["nmr"]
            lines.append("$nmr")
            if nmr["mode"] in ("integration", "mixed") and nmr.get("n_H_list"):
                lines.append(f"integration: {', '.join(str(n) for n in nmr['n_H_list'])}")
            if nmr["mode"] in ("shift", "mixed") and nmr.get("targets"):
                lines.append(f"shift: {', '.join(nmr['targets'])}")
        elif section == "spectra" and parsed.get("spectra") is not None:
            lines.append("$spectra")
            transparent = parsed["spectra"].get("transparent", [])
            if transparent:
                lines.append(f"transparent: {', '.join(transparent)}")
        return lines

    # Walk script line by line: emit comments as-is, replace section content with rebuilt block
    param_lines = []
    _emitted_sections = set()
    _in_section = None
    for _l in script_text.splitlines():
        _stripped = _l.strip()
        if _stripped.startswith('$'):
            _sec = _stripped.lower().lstrip('$').split()[0]
            _in_section = _sec
            if _sec not in _emitted_sections:
                block = _param_block(_sec)
                if block:
                    param_lines.extend(block)
                    _emitted_sections.add(_sec)
        elif _stripped.startswith('#'):
            param_lines.append(_stripped)
        elif _stripped == '' and _in_section is not None:
            param_lines.append('')

    df_parameters = pd.DataFrame({'Parameters': param_lines})

    # ── WRITE & POST-PROCESS ──────────────────────────────────────────────────
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_data.to_excel(writer, sheet_name='data', index=False)
        df_script.to_excel(writer, sheet_name='script', index=False)
        df_parameters.to_excel(writer, sheet_name='parameters', index=False)

        wb = writer.book

        # ── Script tab: paths in B1 / C1, auto column widths ─────────────────
        ws_script = wb['script']
        if script_path:
            ws_script['B1'] = str(script_path)
        if input_path:
            ws_script['C1'] = str(input_path)

        # col A: fit to longest script line
        max_a = max((len(str(ws_script.cell(r, 1).value or ''))
                     for r in range(1, ws_script.max_row + 1)), default=10)
        ws_script.column_dimensions['A'].width = min(max_a + 2, 120)
        # col B / C: fit to path length
        ws_script.column_dimensions['B'].width = min(len(str(script_path or '')) + 2, 120)
        ws_script.column_dimensions['C'].width = min(len(str(input_path  or '')) + 2, 120)

        # ── Parameters tab: auto column width for col A ───────────────────────
        ws_params = wb['parameters']
        max_p = max((len(str(ws_params.cell(r, 1).value or ''))
                     for r in range(1, ws_params.max_row + 1)), default=10)
        ws_params.column_dimensions['A'].width = min(max_p + 2, 120)

    return buffer.getvalue(), filename

# ─────────────────────────────────────────────
# PLOT SNAPSHOT FUNCTIONALITY
# ─────────────────────────────────────────────

def generate_parameters_text(parsed, params, logK_vals):
    """Generate parameters text in script format for snapshot."""
    param_lines = []
    
    # $concentrations section with current values
    param_lines.append("$concentrations")
    for cname, cval in parsed["concentrations"].items():
        current_val = params["conc0"].get(cname[:-1] if cname.endswith("0") else cname, cval)
        param_lines.append(f"{cname} = {current_val:.4f}")
    param_lines.append("")
    
    # $volumes section with current values  
    param_lines.append("$volumes")
    param_lines.append(f"V0 = {params['V0_mL']:.4f}")
    param_lines.append("")
    
    # $titrant section with current values
    if params.get("titrant_is_solid", False):
        param_lines.append("$titrant solid")
        tit_ratios = parsed.get("titrant_ratios", {})
        for tkey in parsed["titrant"].keys():
            param_lines.append(f"{tkey}; {tit_ratios.get(tkey, 1.0):.4f}")
    else:
        param_lines.append("$titrant")
        for tkey, tval in parsed["titrant"].items():
            current_val = params["titrant_mMs"].get(tkey[:-1] if tkey.endswith("t") else tkey, tval)
            param_lines.append(f"{tkey} = {current_val:.4f}")
    param_lines.append("")
    
    # $reactions section with current logK values
    param_lines.append("$reactions")
    for eq in parsed["equilibria"]:
        reactants_str = []
        for coeff, species in eq["reactants"]:
            if coeff > 1:
                reactants_str.append(f"{coeff} {species}")
            else:
                reactants_str.append(species)
        
        # Format products (handle both single and multiple products)
        products_str = []
        if "products" in eq:
            for coeff, species in eq["products"]:
                if coeff > 1:
                    products_str.append(f"{coeff} {species}")
                else:
                    products_str.append(species)
        elif "product" in eq:
            # Backwards compatibility
            prod_coeff, prod_species = eq["product"]
            if prod_coeff > 1:
                products_str.append(f"{prod_coeff} {prod_species}")
            else:
                products_str.append(prod_species)
        
        reactants_part = " + ".join(reactants_str)
        products_part = " + ".join(products_str)
        kname = eq["kname"]
        
        # Get current logK value (may be different from original due to fitting)
        current_logK = logK_vals.get(kname, eq["logK"])
        
        param_lines.append(f"{reactants_part} = {products_part}; log {kname} = {current_logK:.4f}")
    
    if parsed.get("variables"):
        param_lines.append("")
        param_lines.append("$variables") 
        for var_name, expression in parsed["variables"].items():
            param_lines.append(f"{var_name} = {expression}")
    
    if parsed.get("plot_y") or parsed.get("plot_x_expr"):
        param_lines.append("")
        param_lines.append("$plot")
        param_lines.append(f"xmax = {params['maxEquiv']:.4f}")  # xmax first
        if parsed.get("plot_x_expr"):
            param_lines.append(f"x = {parsed['plot_x_expr']}")  # x second
        if parsed.get("plot_y"):
            y_targets = ", ".join(parsed["plot_y"])
            param_lines.append(f"y = {y_targets}")             # y third
    
    return "\n".join(param_lines)


def generate_kinetics_parameters_text(parsed, logk_dict, script_text=""):
    """Generate parameters text in script format for kinetics snapshot."""

    def _block(section):
        lines = []
        if section == "concentrations":
            lines.append("$concentrations")
            for cname, cval in parsed["concentrations"].items():
                lines.append(f"{cname} = {cval:.4f}")
        elif section == "volumes" and parsed.get("volumes"):
            lines.append("$volumes")
            for vname, vval in parsed["volumes"].items():
                lines.append(f"{vname} = {vval:.4f} mL")
        elif section in ("reactions", "equilibria"):
            lines.append("$reactions")
            for eq in parsed["equilibria"]:
                lhs = " + ".join(sp if c == 1 else f"{c} {sp}" for c, sp in eq["reactants"])
                rhs = " + ".join(sp if c == 1 else f"{c} {sp}" for c, sp in eq["products"])
                kval = 10.0 ** logk_dict.get(eq["kname"], eq["logK"])
                lines.append(f"{lhs} = {rhs}; {eq['kname']} = {kval:.4g}")
            for rxn in parsed["kinetics"]:
                lhs = " + ".join(sp if c == 1 else f"{c} {sp}" for c, sp in rxn["reactants"])
                rhs = " + ".join(sp if c == 1 else f"{c} {sp}" for c, sp in rxn["products"])
                arrow = "<>" if rxn["type"] == "reversible_kinetic" else ">"
                k_fwd = 10.0 ** logk_dict.get(rxn["kname"], rxn["log_k"])
                k_str = f"{rxn['kname']} = {k_fwd:.4g}"
                if "krname" in rxn:
                    k_rev = 10.0 ** logk_dict.get(rxn["krname"], rxn["log_kr"])
                    k_str += f"; {rxn['krname']} = {k_rev:.4g}"
                lines.append(f"{lhs} {arrow} {rhs}; {k_str}")
        elif section == "variables" and parsed.get("variables"):
            lines.append("$variables")
            for vname, expr in parsed["variables"].items():
                lines.append(f"{vname} = {expr}")
        elif section == "plot":
            lines.append("$plot")
            lines.append(f"xmax = {parsed['plot_xmax']:.4f}")
            if parsed.get("plot_y"):
                lines.append(f"y = {', '.join(parsed['plot_y'])}")
        elif section == "nmr" and parsed.get("nmr"):
            nmr = parsed["nmr"]
            lines.append("$nmr")
            if nmr["mode"] in ("integration", "mixed") and nmr.get("n_H_list"):
                lines.append(f"integration: {', '.join(str(n) for n in nmr['n_H_list'])}")
            if nmr["mode"] in ("shift", "mixed") and nmr.get("targets"):
                lines.append(f"shift: {', '.join(nmr['targets'])}")
        elif section == "spectra" and parsed.get("spectra") is not None:
            lines.append("$spectra")
            transparent = parsed["spectra"].get("transparent", [])
            if transparent:
                lines.append(f"transparent: {', '.join(transparent)}")
        return lines

    # Fallback: if no script provided, use a sensible default order
    if not script_text.strip():
        result = []
        for sec in ["concentrations", "volumes", "reactions", "variables", "plot", "nmr", "spectra"]:
            block = _block(sec)
            if block:
                if result:
                    result.append("")
                result.extend(block)
        return "\n".join(result)

    # Walk script line by line: emit comments as-is, replace section content with rebuilt block
    result = []
    _emitted_sections = set()
    _in_section = None
    for _l in script_text.splitlines():
        _stripped = _l.strip()
        if _stripped.startswith('$'):
            _sec = _stripped.lower().lstrip('$').split()[0]
            _in_section = _sec
            if _sec not in _emitted_sections:
                block = _block(_sec)
                if block:
                    result.extend(block)
                    _emitted_sections.add(_sec)
        elif _stripped.startswith('#'):
            result.append(_stripped)
        elif _stripped == '' and _in_section is not None:
            result.append('')

    return "\n".join(result)

def text_to_image(text, width=600, font_size=22):
    """Convert text to PIL Image with proper formatting."""
    try:
        # Try to load a monospace font for better alignment
        font = ImageFont.truetype("/System/Library/Fonts/Courier.ttc", font_size)
    except:
        try:
            # Fallback to default font
            font = ImageFont.load_default()
        except:
            font = None
    
    # Split text into lines
    lines = text.split('\n')
    
    # Calculate image dimensions with better spacing
    if font:
        # Get text dimensions using textbbox with sample text
        sample_bbox = ImageDraw.Draw(Image.new('RGB', (1, 1))).textbbox((0, 0), "Mg", font=font)
        line_height = (sample_bbox[3] - sample_bbox[1]) + 8  # Increased padding between lines
    else:
        line_height = font_size + 8
    
    height = len(lines) * line_height + 60  # Increased padding
    
    # Create image with white background
    img = Image.new('RGB', (width, height), color='white')
    draw = ImageDraw.Draw(img)
    
    # Draw text line by line with better formatting
    y = 30  # Increased top margin
    for line in lines:
        # Add visual distinction for section headers
        if line.startswith('$'):
            # Make section headers slightly bolder by drawing with slight offset
            if font:
                draw.text((25, y), line, fill='#333333', font=font)
                draw.text((26, y), line, fill='#333333', font=font)  # Slight offset for bold effect
            else:
                draw.text((25, y), line, fill='#333333')
        else:
            # Regular content with slight indent
            if font:
                draw.text((30, y), line, fill='black', font=font)
            else:
                draw.text((30, y), line, fill='black')
        y += line_height
    
    return img

def create_snapshot(fig, parsed, params, logK_vals):
    """Create side-by-side snapshot of plot and parameters."""
    # Generate timestamp for filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Network_snapshot_{timestamp}.png"
    
    # Convert plotly figure to image
    try:
        # Export figure as PNG bytes
        plot_bytes = fig.to_image(format="png", width=800, height=600, engine="kaleido")
        plot_img = Image.open(io.BytesIO(plot_bytes))
    except Exception as e:
        # If kaleido fails, create a placeholder
        plot_img = Image.new('RGB', (800, 600), color='lightgray')
        draw = ImageDraw.Draw(plot_img)
        draw.text((400, 300), "Plot export failed", fill='black', anchor='mm')
    
    # Generate parameters text
    params_text = generate_parameters_text(parsed, params, logK_vals)
    
    # Convert parameters to image
    params_img = text_to_image(params_text, width=600)
    
    # Resize images to have same height
    target_height = max(plot_img.height, params_img.height)
    
    if plot_img.height != target_height:
        new_width = int(plot_img.width * target_height / plot_img.height)
        plot_img = plot_img.resize((new_width, target_height), Image.Resampling.LANCZOS)
    
    if params_img.height != target_height:
        new_width = int(params_img.width * target_height / params_img.height)
        params_img = params_img.resize((new_width, target_height), Image.Resampling.LANCZOS)
    
    # Combine images side by side
    combined_width = plot_img.width + params_img.width
    combined_img = Image.new('RGB', (combined_width, target_height), color='white')
    
    # Paste images
    combined_img.paste(plot_img, (0, 0))
    combined_img.paste(params_img, (plot_img.width, 0))
    
    # Convert to bytes for download
    buffer = io.BytesIO()
    combined_img.save(buffer, format='PNG')
    
    return buffer.getvalue(), filename

# ─────────────────────────────────────────────

def _export_kinetics_excel(kin_curve, t_vals, plot_y_names, parsed, logk_dict,
                            script_text, variables, script_path=None, input_path=None):
    """Export kinetics simulation results to Excel bytes."""
    from io import BytesIO
    import openpyxl
    from openpyxl.utils import get_column_letter

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"Equinix_{timestamp}.xlsx"

    buf  = BytesIO()
    data = {"Time [s]": t_vals}
    all_sp = _collect_all_kinetic_species(parsed)
    for sp in all_sp:
        if sp in kin_curve:
            data[sp] = kin_curve[sp]
    for vname, expr in variables.items():
        ns = {s: kin_curve.get(s, np.zeros_like(t_vals)) for s in all_sp}
        try:
            data[vname] = np.array(eval(expr, {"__builtins__": {}, "np": np}, ns))
        except Exception:
            pass

    df_data   = pd.DataFrame(data)
    df_script = pd.DataFrame({"Script": script_text.split("\n")})
    params_text = generate_kinetics_parameters_text(parsed, logk_dict, script_text)
    df_params = pd.DataFrame({"Parameters": params_text.split("\n")})

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_data.to_excel(writer, sheet_name="Data", index=False)
        df_script.to_excel(writer, sheet_name="Script", index=False)
        df_params.to_excel(writer, sheet_name="Parameters", index=False)

        wb = writer.book

        # ── Script tab: paths in B1 / C1, auto column widths ─────────────────
        ws_script = wb["Script"]
        if script_path:
            ws_script["B1"] = str(script_path)
        if input_path:
            ws_script["C1"] = str(input_path)

        max_a = max((len(str(ws_script.cell(r, 1).value or ''))
                     for r in range(1, ws_script.max_row + 1)), default=10)
        ws_script.column_dimensions['A'].width = min(max_a + 2, 120)
        ws_script.column_dimensions['B'].width = min(len(str(script_path or '')) + 2, 120)
        ws_script.column_dimensions['C'].width = min(len(str(input_path  or '')) + 2, 120)

        # ── Parameters tab: auto column width for col A ───────────────────────
        ws_params = wb["Parameters"]
        max_p = max((len(str(ws_params.cell(r, 1).value or ''))
                     for r in range(1, ws_params.max_row + 1)), default=10)
        ws_params.column_dimensions['A'].width = min(max_p + 2, 120)

    return buf.getvalue()

def load_experimental_data(file_bytes) -> dict:
    """
    Read experimental data from an Excel file using pandas.

    Liquid mode:
      - First column: volume of titrant added (mL)
      - Subsequent columns: species/variable names

    Solid mode:
      - First column: x-axis values (equivalents or whatever the x expression is)
        The column header is stored under the key "_x_col_header" in the result.
      - Subsequent columns: species/variable names

    Returns dict:
      {column_name: {"v_add_mL": np.ndarray, "y": np.ndarray}}
      plus "_x_col_header": str  (the header of column A)
    where NaN rows are already dropped per column.
    """
    try:
        import io
        df = pd.read_excel(io.BytesIO(file_bytes), engine='openpyxl')
        
        if df.empty:
            return {}
        
        headers = list(df.columns)
        if len(headers) < 2:
            return {}
            
        x_col    = headers[0]
        data_cols = headers[1:]
        
        x_series = pd.to_numeric(df[x_col], errors='coerce')
        if x_series.dropna().empty:
            return {}
        
        result = {"_x_col_header": str(x_col)}
        for col in data_cols:
            try:
                col_series = pd.to_numeric(df[col], errors='coerce')
                valid_mask = x_series.notna() & col_series.notna()
                if valid_mask.sum() > 0:
                    x_valid = x_series[valid_mask].values
                    y_valid = col_series[valid_mask].values
                    finite_mask = np.isfinite(x_valid) & np.isfinite(y_valid)
                    if finite_mask.sum() > 0:
                        result[col] = {
                            "v_add_mL": x_valid[finite_mask],   # mL in liquid, equiv in solid
                            "y":        y_valid[finite_mask]
                        }
            except Exception:
                continue
        
        return result
        
    except Exception:
        return {}


def load_spectra_data(file_bytes) -> dict:
    """
    Read UV-Vis spectral data from an Excel file.

    Format:
      Row 0: col A = label (e.g. "V"), cols B+ = wavelength values (nm, floats)
      Rows 1+: col A = volume of titrant added (mL), cols B+ = absorbance

    Returns:
      {
        "wavelengths": np.ndarray (n_wavelengths,),
        "x_vals":      np.ndarray (n_spectra,),   # volume added in mL
        "A":           np.ndarray (n_spectra, n_wavelengths),
      }
    or {} on failure.
    """
    try:
        import io as _io
        df = pd.read_excel(_io.BytesIO(file_bytes), header=None)
        if df.shape[0] < 2 or df.shape[1] < 2:
            return {}

        wavelengths = pd.to_numeric(pd.Series(df.iloc[0, 1:].values), errors="coerce").values.astype(float)

        x_raw = pd.to_numeric(pd.Series(df.iloc[1:, 0].values), errors="coerce").values.astype(float)
        A_raw = df.iloc[1:, 1:].apply(pd.to_numeric, errors="coerce").values.astype(float)

        # Keep only rows where x and all absorbances are finite
        valid = np.isfinite(x_raw) & np.all(np.isfinite(A_raw), axis=1)
        # Also keep only wavelength columns that are finite
        wl_valid = np.isfinite(wavelengths)

        if valid.sum() == 0 or wl_valid.sum() == 0:
            return {}

        return {
            "wavelengths": wavelengths[wl_valid],
            "x_vals":      x_raw[valid],
            "A":           A_raw[np.ix_(valid, wl_valid)],
        }
    except Exception:
        return {}



def _pub_tight_bounds(plotly_fig):
    """Compute tight data bounds (xmin, xmax, ymin, ymax) from all traces."""
    xs_all, ys_all = [], []
    for trace in plotly_fig.data:
        if trace.x is not None:
            xs_all.extend([v for v in trace.x if v is not None])
        if trace.y is not None:
            ys_all.extend([v for v in trace.y if v is not None])
    if not xs_all or not ys_all:
        return 0.0, 1.0, 0.0, 1.0
    return float(min(xs_all)), float(max(xs_all)), float(min(ys_all)), float(max(ys_all))


def _pub_figure_bytes(plotly_fig, x_label: str = "", y_label: str = "",
                      xlim=None, ylim=None) -> bytes:
    """
    Convert a Plotly figure to a publication-quality matplotlib PNG.
    3 in wide, 600 dpi, Arial 9 pt, inward ticks all four sides, full frame,
    no title, no legend. Axis limits set exactly to xlim/ylim — no padding.
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    import io as _io

    # Explicitly list "Arial" (regular) first — matplotlib will match the
    # correct family on the user's system without falling back to Arial Narrow.
    plt.rcParams.update({
        "pdf.fonttype":      42,   # embed as TrueType → fully editable in Illustrator
        "ps.fonttype":       42,
        "font.family":       "sans-serif",
        "font.sans-serif":   ["Arial", "Helvetica", "DejaVu Sans"],
        "font.size":         9,
        "axes.labelsize":    9,
        "xtick.labelsize":   9,
        "ytick.labelsize":   9,
        "axes.linewidth":    0.8,
        "xtick.major.width": 0.8, "ytick.major.width": 0.8,
        "xtick.minor.width": 0.6, "ytick.minor.width": 0.6,
        "xtick.major.size":  3.5, "ytick.major.size":  3.5,
        "xtick.minor.size":  2.0, "ytick.minor.size":  2.0,
        "xtick.direction":   "in", "ytick.direction":  "in",
        "xtick.top":         True,  "ytick.right":     True,
        "axes.spines.top":   True,  "axes.spines.right": True,
        "figure.dpi":        600,   "savefig.dpi":     600,
    })

    fig, ax = plt.subplots(figsize=(3.0, 2.4))

    # ── Traces ───────────────────────────────────────────────────────────────
    _cc = plt.rcParams["axes.prop_cycle"].by_key()["color"]
    _ci = 0
    for trace in plotly_fig.data:
        xs = list(trace.x) if trace.x is not None else []
        ys = list(trace.y) if trace.y is not None else []
        if not xs or not ys:
            continue
        # colour
        _col = None
        if hasattr(trace, "line") and trace.line and trace.line.color:
            _col = trace.line.color
        elif hasattr(trace, "marker") and trace.marker:
            _mc = getattr(trace.marker, "color", None)
            if isinstance(_mc, str):
                _col = _mc
        if not _col:
            _col = _cc[_ci % len(_cc)]; _ci += 1
        # line width
        lw = 0.8  # match frame line width
        mode = getattr(trace, "mode", "lines") or "lines"
        if "lines" in mode and "markers" in mode:
            ax.plot(xs, ys, color=_col, lw=lw,
                    marker="o", markersize=2.5, markeredgewidth=0)
        elif "markers" in mode:
            ax.scatter(xs, ys, color=_col, s=9, linewidths=0, zorder=3)
        else:
            ax.plot(xs, ys, color=_col, lw=lw)

    # ── Labels ───────────────────────────────────────────────────────────────
    layout = plotly_fig.layout
    if not x_label and layout.xaxis and layout.xaxis.title:
        x_label = layout.xaxis.title.text or ""
    if not y_label and layout.yaxis and layout.yaxis.title:
        y_label = layout.yaxis.title.text or ""

    if x_label:
        ax.set_xlabel(x_label)
    if y_label:
        # Convert unicode superscripts to matplotlib math notation
        y_mpl = (y_label
                 .replace("⁻¹", r"$^{-1}$")   # ⁻¹
                 .replace("⁻²", r"$^{-2}$")   # ⁻²
                 .replace("⁻",       r"$^{-}$"))    # bare ⁻
        ax.set_ylabel(y_mpl)

    # ── Exact axis limits — zero matplotlib margin ────────────────────────────
    ax.margins(0)
    if xlim:
        ax.set_xlim(xlim[0], xlim[1])
    if ylim:
        ax.set_ylim(ylim[0], ylim[1])

    ax.minorticks_on()
    fig.tight_layout(pad=0.4)

    buf = _io.BytesIO()
    fig.savefig(buf, format="pdf", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _pub_axis_range(plotly_fig):
    """
    Return (xlim, ylim) for the publication figure.
    Reads the explicit range from the Plotly layout first (respects xmax set
    by the user); falls back to tight data bounds if the axis is on autorange.
    """
    layout = plotly_fig.layout
    xmn_d, xmx_d, ymn_d, ymx_d = _pub_tight_bounds(plotly_fig)

    # X range
    x_rng = getattr(layout.xaxis, "range", None)
    if x_rng and len(x_rng) == 2 and x_rng[0] is not None and x_rng[1] is not None:
        xlim = (float(x_rng[0]), float(x_rng[1]))
    else:
        xlim = (xmn_d, xmx_d)

    # Y range — always use tight data bounds clipped to the x range
    # (so dots outside xmax don't inflate the y scale)
    ys_in_range = []
    for trace in plotly_fig.data:
        if trace.x is None or trace.y is None:
            continue
        for xi, yi in zip(trace.x, trace.y):
            if xi is not None and yi is not None and xlim[0] <= float(xi) <= xlim[1]:
                ys_in_range.append(float(yi))
    if ys_in_range:
        ylim = (min(ys_in_range), max(ys_in_range))
    else:
        ylim = (ymn_d, ymx_d)

    return xlim, ylim


def _pub_download_button(plotly_fig, key: str,
                          x_label: str = "", y_label: str = "") -> None:
    """Render a publication-figure download button."""
    from datetime import datetime as _dt
    try:
        xlim, ylim = _pub_axis_range(plotly_fig)
        _ts  = _dt.now().strftime("%Y%m%d_%H%M%S")
        _png = _pub_figure_bytes(plotly_fig, x_label, y_label,
                                 xlim=xlim, ylim=ylim)
        st.download_button(
            "📐 Publication figure (pdf)",
            data=_png,
            file_name=f"{key}_{_ts}.pdf",
            mime="application/pdf",
            key=f"_pubfig_{key}",
        )
    except Exception as _e:
        st.caption(f"Export failed: {_e}")




def _plot_backcalc_dots(fig, c_bc_pairs: dict, plot_y_names: list,
                        variables: dict, all_species: list, trace_colors: dict,
                        label_suffix: str = "(NMR)"):
    """
    Add back-calculated concentration dots to `fig` for every quantity in
    `plot_y_names`.

    Parameters
    ----------
    c_bc_pairs  : {sp: (x_arr, c_arr)}  — one entry per back-calculated species.
                  All entries share the same x_arr (or close enough to interpolate).
    plot_y_names: the list from $plot y = ...
    variables   : parsed["variables"]
    all_species : all species in the model
    trace_colors: {name: hex_color}
    label_suffix: appended to legend name (e.g. "(NMR)")

    Logic
    -----
    For each name in plot_y_names:
      • If it is a back-calc species  → plot directly.
      • If it is a $variable          → evaluate the expression from the
                                        back-calc concentrations and plot.
      • Otherwise                     → skip.
    This correctly handles cases like %GH = GH/(G+GH) where the user plots a
    derived quantity rather than a raw concentration.
    """
    if not c_bc_pairs:
        return

    # Shared x-axis: use the first species' x_arr as the reference grid.
    ref_sp   = next(iter(c_bc_pairs))
    x_ref, _ = c_bc_pairs[ref_sp]
    n_pts    = len(x_ref)

    # Build {sp: c_arr_on_x_ref} — interpolate if necessary.
    c_on_ref = {}
    for sp, (x_sp, c_sp) in c_bc_pairs.items():
        if len(x_sp) == n_pts and np.allclose(x_sp, x_ref):
            c_on_ref[sp] = c_sp
        else:
            c_on_ref[sp] = np.interp(x_ref, x_sp, c_sp)

    for name in plot_y_names:
        if name in c_on_ref:
            y_arr = c_on_ref[name]
        elif name in variables:
            # Evaluate $variable expression pointwise from back-calc concentrations
            fake_curve   = {sp: c_on_ref.get(sp, np.zeros(n_pts)) for sp in all_species}
            fake_network = {"all_species": all_species}
            y_arr = compute_variable_curve(name, variables, fake_curve,
                                           fake_network, x_ref)
        else:
            continue   # not representable from back-calc data

        color = trace_colors.get(name, "#AAAAAA")
        fig.add_trace(go.Scatter(
            x=x_ref, y=y_arr,
            mode="markers", name=f"{name} {label_suffix}",
            marker=dict(color=color, size=6, symbol="circle"),
            showlegend=True,
        ))

def _infer_unit(name, variables, species_set, _seen=None):
    """
    Infer the unit of a variable or species.
    Returns: "mM", "fraction", or "unknown".
    Species concentrations → "mM".
    Ratios of same-unit expressions → "fraction".
    Sums/differences of mM terms → "mM".
    """
    if _seen is None:
        _seen = set()
    if name in _seen:
        return "unknown"   # circular
    _seen = _seen | {name}

    if name in species_set:
        return "mM"

    if name not in variables:
        return "unknown"

    expr = variables[name].strip()

    # Try to detect division at the top level (not inside parens)
    depth = 0
    div_pos = -1
    for i, ch in enumerate(expr):
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == '/' and depth == 0:
            div_pos = i
            break

    if div_pos != -1:
        # ratio: infer units of numerator and denominator
        numer = expr[:div_pos].strip()
        denom = expr[div_pos+1:].strip()
        # strip outer parens
        def strip_parens(s):
            s = s.strip()
            while s.startswith('(') and s.endswith(')'):
                s = s[1:-1].strip()
            return s
        numer = strip_parens(numer)
        denom = strip_parens(denom)
        # Get units of both sides (use first token as representative)
        import re
        numer_names = re.findall(r'[A-Za-z_][A-Za-z0-9_]*', numer)
        denom_names = re.findall(r'[A-Za-z_][A-Za-z0-9_]*', denom)
        nu = set(_infer_unit(n, variables, species_set, _seen) for n in numer_names) if numer_names else {"number"}
        du = set(_infer_unit(n, variables, species_set, _seen) for n in denom_names) if denom_names else {"number"}
        # If both sides are mM (same unit) → dimensionless fraction
        if nu <= {"mM"} and du <= {"mM"}:
            return "fraction"
        # If numerator is fraction and denominator is number/fraction → fraction
        if nu <= {"fraction", "number"} and du <= {"fraction", "number"}:
            return "fraction"
        return "unknown"

    # No top-level division → sum/difference of terms
    import re
    names = re.findall(r'[A-Za-z_][A-Za-z0-9_]*', expr)
    units = set(_infer_unit(n, variables, species_set, _seen) for n in names) if names else set()
    if units <= {"mM"}:
        return "mM"
    if units <= {"fraction"}:
        return "fraction"
    return "unknown"


def _infer_y_label(plot_y_names, parsed, network):
    """Return appropriate y-axis label based on the units of the plotted quantities."""
    variables   = parsed.get("variables", {})
    all_species = set(network.get("all_species", []))

    units = set()
    for name in plot_y_names:
        u = _infer_unit(name, variables, all_species)
        units.add(u)

    if units <= {"fraction"}:
        return "Fraction"
    if units <= {"mM"}:
        return "Concentration [mM]"
    if "fraction" in units and "mM" in units:
        return "Concentration [mM] / Fraction"
    return "Concentration [mM]"


